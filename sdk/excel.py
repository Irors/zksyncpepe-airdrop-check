import openpyxl
from openpyxl.styles import Border, Side, Alignment, Font

class Excel:

    workbook = openpyxl.Workbook()
    sheet = workbook.active

    sheet['A1'] = 'Address'
    sheet['B1'] = 'Amount'


    # Настройка размера столбца A
    column = sheet.column_dimensions[openpyxl.utils.get_column_letter(1)]
    column.width = 16

    # Настройка размера столбца B
    column_B = sheet.column_dimensions[openpyxl.utils.get_column_letter(2)]
    column_B.width = 16

    # Создаем объект для задания стиля выравнивания и отступов
    alignment_style = Alignment(horizontal='center', vertical='center', indent=3, wrap_text=True)

    # Создаем объект для задания стиля границ
    border_style = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Делаем все ячейки с жирным шрифтом
    for i in ('AB'):
        cell = sheet[f'{i}1']
        bold_font = Font(bold=True)
        cell.font = bold_font

        # Применяем стиль выравнивания и отступов к ячейке
        cell.alignment = alignment_style
        # Применяем стиль границ
        cell.border = border_style
